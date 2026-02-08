from fastapi import FastAPI, Depends, HTTPException, BackgroundTasks
from sqlalchemy.orm import Session
from typing import List, Dict, Any
from fastapi.middleware.cors import CORSMiddleware
import os
import secrets
from datetime import datetime, timedelta
from dotenv import load_dotenv
import yaml
import openpyxl
import httpx
from openpyxl.utils import get_column_letter
# Load environment variables from .env file
load_dotenv()

print(f"DEBUG: GCS_BUCKET_NAME={os.getenv('GCS_BUCKET_NAME')}")
print(f"DEBUG: GOOGLE_APPLICATION_CREDENTIALS={os.getenv('GOOGLE_APPLICATION_CREDENTIALS')}")

from app.db import models, session
from app.schemas import brief as brief_schema
from app.schemas import submission as plan_schema # Renamed for clarity
from app.schemas import user as user_schema
from app.core import security, gcs, exceptions

from contextlib import asynccontextmanager

@asynccontextmanager
async def lifespan(app: FastAPI):
    # This runs when the server starts
    print("🚀 Server starting up...")
    yield
    # This runs when the server stops
    print("🛑 Server shutting down.")

app = FastAPI(title="Brief Ecosystem - Production API", lifespan=lifespan)


 
# CORS - Add this block RIGHT AFTER app = FastAPI()
# CORS - Add this block RIGHT AFTER app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=False, allow_methods=["*"], allow_headers=["*"])

# Register Global Exception Handler
app.add_exception_handler(exceptions.BriefAppException, exceptions.global_exception_handler)
 
@app.get("/health")
async def health():
    return {"status": "healthy"}

@app.get("/")
def read_root():
    return {"status": "online", "version": "2.0.0-relational"}

@app.post("/login", response_model=user_schema.LoginResponse)
def login(payload: user_schema.LoginRequest, db: Session = Depends(session.get_db)):
    """
    Real DB Login: Validates against 'users' table and generates a session token.
    1. Validates password.
    2. Deactivates all previous active tokens for the user (Single Session).
    3. Generates a new Token with 10-day expiry.
    """
    
    user_record = db.query(models.User).filter(models.User.email == payload.email).first()
    
    # Real Hashing Verification
    if not user_record or not security.verify_password(payload.password, user_record.password):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    # 2. Single Session Check: Deactivate all old tokens for this user
    db.query(models.Token).filter(
        models.Token.user_id == user_record.id,
        models.Token.is_active == True
    ).update({"is_active": False})
    
    # 3. Create New Token
    new_token_str = secrets.token_hex(32)
    expires_at = models.get_utc_now() + timedelta(days=10)
    
    new_token_record = models.Token(
        user_id=user_record.id,
        access_token=new_token_str,
        expires_at=expires_at,
        is_active=True
    )
    db.add(new_token_record)
    db.commit()
    
    return {
        "token": new_token_str,
        "user": {
            "email": user_record.email,
            "name": user_record.name,
            "role": user_record.role,
            "agencyName": user_record.agency.name if user_record.agency else None,
            "id": user_record.id
        }
    }

# --- MANAGEMENT ENDPOINTS ---

@app.get("/agencies", response_model=List[dict])
def list_agencies(db: Session = Depends(session.get_db), current_user: dict = Depends(security.get_current_user)):
    """Returns a list of all agencies."""
    agencies = db.query(models.Agency).all()
    return [{"id": a.id, "name": a.name} for a in agencies]

@app.get("/clients", response_model=List[dict])
def list_clients(db: Session = Depends(session.get_db), current_user: dict = Depends(security.get_current_user)):
    """Returns a list of clients."""
    clients = db.query(models.Client).all()
    return [{"id": c.id, "name": c.name} for c in clients]

# --- BRIEF WORKFLOW ---

@app.post("/briefs", response_model=brief_schema.BriefResponse)
def create_brief(
    brief: brief_schema.BriefCreate,
    db: Session = Depends(session.get_db),
    current_user: dict = Depends(security.verify_ds_group)
):
    """
    DS GROUP ONLY: Creates a new Brief.
    """
    # 1. Create the Brief record
    db_brief = models.Brief(
        client_id=current_user["client_id"],
        brand_name=brief.brand_name,
        division=brief.division,
        creative_name=brief.creative_name,
        objective=brief.objective,
        brief_type=brief.brief_type,
        total_budget=brief.total_budget,
        start_date=brief.start_date,
        end_date=brief.end_date,
        created_by=current_user["id"],
        updated_by=current_user["id"],
        
        # Demographic/Market fields
        demographics_age=brief.demographics_age,
        demographics_gender=brief.demographics_gender,
        demographics_nccs=brief.demographics_nccs,
        demographics_etc=brief.demographics_etc,
        psychographics=brief.psychographics,
        key_markets=brief.key_markets,
        p1_markets=brief.p1_markets,
        p2_markets=brief.p2_markets,
        edit_durations=brief.edit_durations,
        acd=brief.acd,
        dispersion=brief.dispersion,
        advertisement_link=brief.advertisement_link,
        creative_languages=brief.creative_languages,
        scheduling_preference=brief.scheduling_preference,
        miscellaneous=brief.miscellaneous,
        remarks=brief.remarks
        # target_agency_ids removed from DB model
    )
    
    db.add(db_brief)
    db.flush()
    
    # --- AUTO-CREATE AGENCY PLAN SLOTS ---
    # We loop through the IDs provided in the request
    for target_id in brief.target_agency_ids:
        agency = db.query(models.Agency).filter(models.Agency.id == target_id).first()
        if agency:
            new_plan = models.AgencyPlan(
                brief_id=db_brief.id,
                agency_id=agency.id,
                status="DRAFT",
                created_by=current_user["id"],
                updated_by=current_user["id"]
            )
            db.add(new_plan)
            db.flush() # Get plan ID
            
            # Initial Audit
            slot_history = models.HistoryTrail(
                agency_plan_id=new_plan.id,
                action="NEW_BRIEF_CREATED",
                user_id=current_user["id"],
                details=f"The brief was created and assigned to {agency.name}.",
                comment=None
            )
            db.add(slot_history)

    db.commit()
    db.refresh(db_brief)
    
    # Fetch names for the response
    # Fetch names for the response
    target_obs = []
    # In-memory assignment for the response model (not DB persistence)
    db_brief.target_agency_ids = brief.target_agency_ids 
    
    if brief.target_agency_ids:
        agencies = db.query(models.Agency).filter(models.Agency.id.in_(brief.target_agency_ids)).all()
        target_obs = [{"id": a.id, "name": a.name} for a in agencies]
        
    db_brief.target_agency_ids = target_obs # Inject objects
    return db_brief

@app.get("/briefs", response_model=List[brief_schema.BriefResponse])
def list_briefs(
    db: Session = Depends(session.get_db),
    current_user: dict = Depends(security.get_current_user)
):
    """
    Returns full detailed briefs.
    """
    if current_user["role"] == "DS_GROUP":
        db_briefs = db.query(models.Brief).all()
    else:
        db_briefs = db.query(models.Brief).join(models.AgencyPlan).filter(
            models.AgencyPlan.agency_id == current_user["agency_id"]
        ).all()
        
    results = []
    for b in db_briefs:
        # Filter agency plans based on role
        filtered_plans = []
        for p in b.agency_plans:
            if current_user["role"] == "DS_GROUP" or p.agency_id == current_user["agency_id"]:
                
                # Convert to Pydantic Model explicitly to ensure fields are populated
                p_model = plan_schema.AgencyPlanSummary.model_validate(p)
                
                # User Request: Dynamic URL Fallback (Validated -> Flat -> Raw)
                # If a new raw file is uploaded, Step 1 (get_upload_url) resets the paths, so this logic stays accurate.
                latest_blob = p.validated_column_file_path or p.flat_file_path or p.raw_file_path
                if latest_blob:
                    p_model.plan_file_url = gcs.get_signed_url(latest_blob, method="GET", expiration_minutes=8640) # 6 days limit (GCS max is 7 days)
                else:
                    p_model.plan_file_url = None
                
                if not p.plan_file_name:
                    p_model.plan_file_name = "Plan"

                filtered_plans.append(p_model)
        
        # Privacy: Agencies should only see their own ID/Name in the target list
        # We RECONSTRUCT target_agency_ids from the existing plans
        # Because we deleted the JSON column.
        
        # 1. Get all Agency IDs attached to this brief
        all_target_ids = [p.agency_id for p in b.agency_plans]
        
        visible_ids = all_target_ids
        if current_user["role"] == "AGENCY":
            visible_ids = [aid for aid in all_target_ids if aid == current_user["agency_id"]]
            
        # Transform IDs to Objects {id, name}
        target_objects = []
        if visible_ids:
            agencies = db.query(models.Agency).filter(models.Agency.id.in_(visible_ids)).all()
            target_objects = [{"id": a.id, "name": a.name} for a in agencies]

        # We manually construct a dictionary for the response
        brief_data = {
            "id": b.id,
            "status": b.status,
            "brandName": b.brand_name,
            "division": b.division,
            "creativeName": b.creative_name,
            "campaignObjective": b.objective,
            "type": b.brief_type,
            "totalBudget": b.total_budget,
            "startDate": b.start_date,
            "endDate": b.end_date,
            "targetAgencies": target_objects, # Now returning Objects!
            "demographicsAge": b.demographics_age,
            "demographicsGender": b.demographics_gender,
            "demographicsNccs": b.demographics_nccs,
            "demographicsEtc": b.demographics_etc,
            "psychographics": b.psychographics,
            "keyMarkets": b.key_markets,
            "p1Markets": b.p1_markets,
            "p2Markets": b.p2_markets,
            "editDurations": b.edit_durations,
            "acd": b.acd,
            "dispersion": b.dispersion,
            "advertisementLink": b.advertisement_link,
            "creativeLanguages": b.creative_languages,
            "schedulingPreference": b.scheduling_preference,
            "miscellaneous": b.miscellaneous,
            "remarks": b.remarks,
            "createdAt": b.created_at,
            "updatedAt": b.updated_at,
            "creator": b.creator,
            "updater": b.updater,
            "agencyPlans": filtered_plans
        }
        results.append(brief_data)
        
    return results

@app.get("/briefs/{brief_id}", response_model=brief_schema.BriefFullDetail)
def get_brief_detail(
    brief_id: int,
    db: Session = Depends(session.get_db),
    current_user: dict = Depends(security.get_current_user)
):
    """Returns the same full detail as the list view for consistency."""
    db_brief = db.query(models.Brief).filter(models.Brief.id == brief_id).first()
    if not db_brief:
        raise HTTPException(status_code=404, detail="Brief not found")
        
    if current_user["role"] == "AGENCY":
        has_plan = db.query(models.AgencyPlan).filter(
            models.AgencyPlan.brief_id == brief_id,
            models.AgencyPlan.agency_id == current_user["agency_id"]
        ).first()
        if not has_plan:
            raise HTTPException(status_code=403, detail="Access denied")

    # Filter plans for the response
    filtered_plans = []
    for p in db_brief.agency_plans:
        if current_user["role"] == "DS_GROUP" or p.agency_id == current_user["agency_id"]:
            
            # Convert to Pydantic Model explicitly
            p_model = plan_schema.AgencyPlanSummary.model_validate(p)

            # User Request: Dynamic URL Fallback (Validated -> Flat -> Raw)
            latest_blob = p.validated_column_file_path or p.flat_file_path or p.raw_file_path
            if latest_blob:
                p_model.plan_file_url = gcs.get_signed_url(latest_blob, method="GET", expiration_minutes=8640) # 6 days limit (GCS max is 7 days)
            else:
                p_model.plan_file_url = None
            
            if not p.plan_file_name:
                p_model.plan_file_name = "Plan"
            
            filtered_plans.append(p_model)
    
    # Privacy: Agencies should only see their own ID in the target list
    # Reconstruct from plans
    all_target_ids = [p.agency_id for p in db_brief.agency_plans]
    
    visible_target_agency_ids = all_target_ids
    if current_user["role"] == "AGENCY":
        visible_target_agency_ids = [aid for aid in all_target_ids if aid == current_user["agency_id"]]
        
    # Transform IDs to Objects
    target_objects = []
    if visible_target_agency_ids:
        agencies = db.query(models.Agency).filter(models.Agency.id.in_(visible_target_agency_ids)).all()
        target_objects = [{"id": a.id, "name": a.name} for a in agencies]

    # Construct total response object
    return {
        "id": db_brief.id,
        "status": db_brief.status,
        "brandName": db_brief.brand_name,
        "division": db_brief.division,
        "creativeName": db_brief.creative_name,
        "campaignObjective": db_brief.objective,
        "type": db_brief.brief_type,
        "totalBudget": db_brief.total_budget,
        "startDate": db_brief.start_date,
        "endDate": db_brief.end_date,
        "targetAgencies": target_objects,
        "demographicsAge": db_brief.demographics_age,
        "demographicsGender": db_brief.demographics_gender,
        "demographicsNccs": db_brief.demographics_nccs,
        "demographicsEtc": db_brief.demographics_etc,
        "psychographics": db_brief.psychographics,
        "keyMarkets": db_brief.key_markets,
        "p1Markets": db_brief.p1_markets,
        "p2Markets": db_brief.p2_markets,
        "editDurations": db_brief.edit_durations,
        "acd": db_brief.acd,
        "dispersion": db_brief.dispersion,
        "advertisementLink": db_brief.advertisement_link,
        "creativeLanguages": db_brief.creative_languages,
        "schedulingPreference": db_brief.scheduling_preference,
        "miscellaneous": db_brief.miscellaneous,
        "remarks": db_brief.remarks,
        "createdAt": db_brief.created_at,
        "updatedAt": db_brief.updated_at,
        "creator": db_brief.creator,
        "updater": db_brief.updater,
        "agencyPlans": filtered_plans
    }

# --- GCS UPLOAD HANDSHAKE ---

# --- AGENCY PLAN WORKFLOW (GCS HANDSHAKE) ---

@app.get("/briefs/{brief_id}/plans/{plan_id}/upload-url", response_model=plan_schema.UploadUrlResponse)
def get_upload_url(
    brief_id: int,
    plan_id: int,
    db: Session = Depends(session.get_db),
    current_user: dict = Depends(security.verify_agency)
):
    """
    Step 1: Get the Presigned URL for uploading the RAW file.
    """
    # 1. Verify Plan Ownership
    plan = db.query(models.AgencyPlan).filter(
        models.AgencyPlan.id == plan_id,
        models.AgencyPlan.brief_id == brief_id,
        models.AgencyPlan.agency_id == current_user["agency_id"]
    ).first()
    
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found or access denied.")

    # 2. Construct Path: brief_media_files/brief_id/plan_id/raw/plan.xlsx
    upload_path = f"brief_media_files/{brief_id}/{plan.id}/raw/plan.xlsx"
    
    # 3. Generate Signed PUT URL
    upload_url = gcs.get_signed_url(upload_path, method="PUT", expiration_minutes=8640, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    # 4. PERSIST the intended raw path and RESET processing states
    # This ensures that if a user re-uploads, the latest URL points to the new RAW file.
    plan.raw_file_path = upload_path
    plan.flat_file_path = None
    plan.validated_column_file_path = None
    plan.ai_mappings = None
    plan.human_mappings = None
    db.commit()

    return {
        "uploadUrl": upload_url,
        "planId": plan.id,
        "expiresIn": "15 minutes"
    }


def _external_validation_service_mock(raw_path: str, brief_id: int, plan_id: int):
    """
    Simulates the External Validation Service (Validation Logic).
    - It now physically copies the RAW file to the FLAT path in GCS 
      so downstream testing (Step 3) works with real files.
    """
    # ... (existing mock implementation - we might deprecate this but keeping for now if needed)
    pass

# --- ASSISTANT: SENIOR API INTEGRATION ---
def _call_senior_api_extract(gcs_path: str):
    """
    Calls the Real Senior API 'map-columns' command.
    """
    url = "https://mediaflow-engine-702075889747.asia-south2.run.app/command"
    payload = {
        "command": "map-columns",
        "params": {
            "file_path": gcs_path
        }
    }
    
    print(f"DEBUG: Calling Real Senior API for {gcs_path}")
    
    try:
        # 5 minutes timeout (300s) as requested for external AI calls
        response = httpx.post(url, json=payload, timeout=600.0)  # 10 minutes
        
        if response.status_code != 200:
            raise HTTPException(status_code=response.status_code, detail=f"Senior API Error: {response.text}")
            
        resp_json = response.json()
        if not resp_json.get("success"):
            raise HTTPException(status_code=500, detail=f"Senior API Logic Failure: {resp_json.get('message')}")
            
        return resp_json["data"]
        
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="Senior API Request Timeout")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Senior API Connection Failed: {str(e)}")

def _transform_senior_api_response(agent_resp: dict):
    """
    Adapts the detailed Senior API response to the format needed by our backend.
    """
    ai_mappings = {}
    
    # Process 'mappings' (Successfully mapped columns)
    for c in agent_resp.get("mappings", []):
         idx = c["source_column_index"]
         target = c["target_field"]
         # We map Index -> Standard Name
         ai_mappings[idx] = target

    # Process 'unmapped_source_columns' (We keep them but mapped to "")
    # UPDATE: User requested NOT to map them to empty strings.
    # If we skip them here, _process_excel_extract won't rename them, 
    # so they will keep their ORIGINAL headers from the raw file.
    
    # for c in agent_resp.get("unmapped_source_columns", []):
    #     idx = c["source_column_index"]
    #     target = c["target_field"] # This is "" in the JSON
    #     ai_mappings[idx] = target

    return {
        "ai_mappings": ai_mappings
    }

# --- HELPER: OPENPYXL TRANSFORMATION ---
def _process_excel_extract(local_input_path: str, local_output_path: str, header_row: int, mappings: Dict[int, str]):
    """
    Uses OpenPyXL to:
    1. Delete rows before the header_row (so header becomes Row 1).
    2. Rename headers in the new Row 1 based on mappings.
    3. Preserve formulas and styles.
    """
    wb = openpyxl.load_workbook(local_input_path)
    ws = wb.active
    
    # 1. Crop Rows: Move data UP instead of deleting rows (to preserve formulas)
    if header_row > 1:
        offset = header_row - 1
        max_row = ws.max_row
        max_col = ws.max_column
        max_col_letter = get_column_letter(max_col)
        
        # Range to move: From [header_row, 1] to [max_row, max_col]
        move_ref = f"A{header_row}:{max_col_letter}{max_row}"
        
        # Move UP by 'offset' rows
        ws.move_range(move_ref, rows=-offset, translate=True)
        
        # Now delete the empty rows at the bottom
        ws.delete_rows(max_row - offset + 1, amount=offset)
        
    # 2. Rename Headers (Now at Row 1)
    for col_idx, new_name in mappings.items():
        cell = ws.cell(row=1, column=col_idx + 1) 
        cell.value = new_name
        
    wb.save(local_output_path)
    return True

# --- HELPER: READ SCHEMA YAML ---
def _get_schema_columns():
    """Reads mandatory/optional columns from GCS configs bucket or local fallback."""
    # User's Remote GCS Path
    remote_bucket = "mediaflow-configs"
    remote_blob = "schemas/plan.yaml"
    local_temp_path = "tmp/remote_schema.yaml"
    
    try:
        # 1. Try fetching from GCS (Securely using Service Account)
        gcs.download_from_bucket(remote_bucket, remote_blob, local_temp_path)
        with open(local_temp_path, "r") as f:
            data = yaml.safe_load(f)
            print(f"DEBUG: Successfully loaded schema from gs://{remote_bucket}/{remote_blob}")
    except Exception as e:
        print(f"DEBUG: Remote schema fetch failed ({e}). Falling back to local.")
        # 2. Fallback to local file
        try:
            with open("schemas_plan.yaml", "r") as f:
                data = yaml.safe_load(f)
        except Exception as local_e:
            print(f"ERROR: Local schema fallback also failed: {local_e}")
            return [], []

    mandatory = [c["name"] for c in data.get("columns", {}).get("mandatory", [])]
    optional = [c["name"] for c in data.get("columns", {}).get("optional", [])]
    return mandatory, optional

@app.get("/briefs/{brief_id}/plans/{plan_id}/extract-columns", response_model=plan_schema.ValidateColumnsResponse)
def extract_columns(
    brief_id: int,
    plan_id: int,
    db: Session = Depends(session.get_db),
    current_user: dict = Depends(security.verify_agency)
):
    """
    Step 2: Extract Columns (Replaces Validate Columns)
    - Mocks Senior API to parse headers.
    - Transforms file (Crop + Rename) maintaining formulas.
    - Returns AI mappings and Flat File URL.
    """
    # 1. Verify Plan
    plan = db.query(models.AgencyPlan).filter(
        models.AgencyPlan.id == plan_id,
        models.AgencyPlan.brief_id == brief_id,
        models.AgencyPlan.agency_id == current_user["agency_id"]
    ).first()
    
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")

    # 2. Define Paths
    # Input: brief_media_files/{brief_id}/{plan_id}/raw/plan.xlsx
    # But wait, previous step used: {brief_id}/{plan_id}/raw/plan.xlsx
    # The user asked to CHANGE the path prefix in the implementation plan.
    # We should support the file wherever it was uploaded.
    # For now, let's assume the upload URL generation uses the NEW prefix 'brief_media_files'
    # IF we haven't updated upload-url code yet, we need to handle that.
    # Let's start using the NEW prefix structure for this flow.
    
    raw_blob_name = f"brief_media_files/{brief_id}/{plan_id}/raw/plan.xlsx"
    flat_blob_name = f"brief_media_files/{brief_id}/{plan_id}/flat/plan_flat.xlsx"
    
    local_raw_path = f"tmp/{brief_id}_{plan_id}_raw.xlsx"
    local_flat_path = f"tmp/{brief_id}_{plan_id}_flat.xlsx"
    
    # Ensure tmp directory exists
    os.makedirs("tmp", exist_ok=True)

    # 3. Download Raw
    try:
        gcs.download_file(raw_blob_name, local_raw_path)
    except Exception:
        # Fallback for older upload path without prefix, just in case? 
        # Or better, just Try the old path if new fails for backward compatibility during dev.
        raw_blob_name_old = f"{brief_id}/{plan.id}/raw/plan.xlsx"
        try:
            gcs.download_file(raw_blob_name_old, local_raw_path)
            raw_blob_name = raw_blob_name_old # Update reference
        except Exception as e:
            raise HTTPException(status_code=404, detail="Raw file not found. Please upload first.")

    # 4. Call Real Senior API
    senior_data = _call_senior_api_extract(f"gs://{gcs.BUCKET_NAME}/{raw_blob_name}")
    
    # 5. Transform Response Format
    transformed_resp = _transform_senior_api_response(senior_data)
    
    # User Request: Direct logic based on Senior API response.
    # header_row = data_start_row - 1
    data_start_row = senior_data.get("data_start_row", 1)
    header_row = data_start_row - 1
    if header_row < 1: header_row = 1 
    
    ai_mappings = transformed_resp["ai_mappings"] # {0: "name", ...}

    # 6. Strict Validation: Check if AI mappings cover all mandatory columns
    # We fetch the schema columns from the GCS URL/Local fallback
    mandatory_cols, optional_cols = _get_schema_columns()
    mapped_target_fields = set(ai_mappings.values())
    
    # Rejection Logic: If the number of mapped standard columns is less than mandatory columns
    # OR if specific mandatory columns are missing.
    missing_mandatory = [col for col in mandatory_cols if col not in mapped_target_fields]
    
    if len(mapped_target_fields) < len(mandatory_cols) or missing_mandatory:
        raise HTTPException(
            status_code=400, 
            detail={
                "error": "Incomplete Mapping",
                "message": f"AI was only able to map {len(mapped_target_fields)} standard columns, but {len(mandatory_cols)} are mandatory.",
                "missing_columns": missing_mandatory
            }
        )

    # 7. Transform File (OpenPyXL)
    try:
        _process_excel_extract(local_raw_path, local_flat_path, header_row, ai_mappings)
    except Exception as e:
        print(f"Excel processing failed: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to process Excel file: {str(e)}")
        
    # 6. Upload Flat File
    try:
        gcs.upload_file(local_flat_path, flat_blob_name, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to upload flat file: {str(e)}")

    # 7. Update DB
    plan.flat_file_path = flat_blob_name
    # Pydantic/SQLAlchemy expects keys as strings for JSON usually, let's ensure compatibility
    # ai_mappings keys are ints (0, 1, 2). JSON supports string keys primarily.
    # Convert keys to string for DB storage to be safe
    db_mappings = {str(k): v for k, v in ai_mappings.items()}
    plan.ai_mappings = db_mappings
    db.commit()
    
    # 8. Get Schema Columns
    mandatory, optional = _get_schema_columns()
    
    # 9. Generate Signed URL for Flat File
    flat_url = gcs.get_signed_url(flat_blob_name, method="GET", expiration_minutes=8640)  # 6 days
    
    # 10. Return Response
    # Convert int keys back to string for Pydantic response if schema says Dict[str, str]
    response_mappings = {str(k): v for k, v in ai_mappings.items()} 
    
    # Clean up local files
    if os.path.exists(local_raw_path): os.remove(local_raw_path)
    if os.path.exists(local_flat_path): os.remove(local_flat_path)

    return {
        "flatFileUrl": flat_url,
        "aiMappings": response_mappings,
        "requiredColumns": mandatory,
        "optionalColumns": optional
    }

@app.post("/briefs/{brief_id}/plans/{plan_id}/validate-columns", response_model=plan_schema.ValidateColumnsResponse)
def validate_columns(
    brief_id: int,
    plan_id: int,
    db: Session = Depends(session.get_db),
    current_user: dict = Depends(security.verify_agency)
):
    """
    Step 2: Trigger Senior API to validate raw file and generate mappings.
    """
    plan = db.query(models.AgencyPlan).filter(
        models.AgencyPlan.id == plan_id,
        models.AgencyPlan.brief_id == brief_id,
        models.AgencyPlan.agency_id == current_user["agency_id"]
    ).first()
    
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found.")

    # 1. Check if RAW file exists (skipped for now, assumed flow)
    raw_path = f"{brief_id}/{plan.id}/raw/plan.xlsx"
    
    # 2. Call External Validation Service (Mock)
    validation_result = _external_validation_service_mock(raw_path, brief_id, plan.id)
    
    # 3. STRICT VALIDATION: Check if all REQUIRED columns are present in the AI Mappings
    # (i.e., did the AI find a match for every required column?)
    # or simple check if the FILE has the potential to map to them. 
    # For this mock, we assume 'ai_mappings' keys are the Standard Columns found.
    
    found_columns = validation_result["ai_mappings"].keys() # keys are the Standard Names
    missing_required = [col for col in validation_result["required_columns"] if col not in found_columns]
    
    if missing_required:
        raise HTTPException(
            status_code=400, 
            detail=f"Validation Failed: Missing required columns: {', '.join(missing_required)}"
        )
    
    # 4. Update DB
    plan.flat_file_path = validation_result["flat_path"]
    plan.ai_mappings = validation_result["ai_mappings"]
    plan.raw_file_path = f"brief_media_files/{brief_id}/{plan.id}/raw/plan.xlsx" # Consistent prefix
    plan.updated_at = models.get_utc_now()
    plan.updated_by = current_user["id"]
    
    # 5. Log History
    history = models.HistoryTrail(
        agency_plan_id=plan.id,
        action="COLUMNS_VALIDATED",
        user_id=current_user["id"],
        details="AI Validation completed. Columns mapped."
    )
    db.add(history)
    db.commit()
    
    # 6. Generate View URL
    flat_url = gcs.get_signed_url(plan.flat_file_path, method="GET", expiration_minutes=8640)  # 6 days
    
    return {
        "flatFileUrl": flat_url,
        "aiMappings": plan.ai_mappings,
        "requiredColumns": validation_result["required_columns"],
        "optionalColumns": validation_result["optional_columns"]
    }

@app.post("/briefs/{brief_id}/plans/{plan_id}/update-columns", response_model=plan_schema.UpdateColumnsResponse)
def update_columns(
    brief_id: int,
    plan_id: int,
    payload: plan_schema.UpdateColumnsRequest,
    db: Session = Depends(session.get_db),
    current_user: dict = Depends(security.verify_agency)
):
    """
    Step 3: Confirm mappings, rename columns, and finalize the file.
    """
    plan = db.query(models.AgencyPlan).filter(
        models.AgencyPlan.id == plan_id,
        models.AgencyPlan.brief_id == brief_id,
        models.AgencyPlan.agency_id == current_user["agency_id"]
    ).first()
    
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found.")

    # 1. Define Paths Dynamically
    # UPDATE: We now download the FLAT file (output of Extract) instead of the RAW file.
    flat_path = plan.flat_file_path or f"brief_media_files/{brief_id}/{plan.id}/flat/plan_flat.xlsx"
    local_flat = f"tmp/update_flat_{plan.id}.xlsx"
    local_validated = f"tmp/update_validated_{plan.id}.xlsx"
    os.makedirs("tmp", exist_ok=True)

    # 2. Merge Mappings
    # The payload 'human_mappings' is likely { "0": "programme", ... }
    final_mappings = (plan.ai_mappings or {}).copy()
    for k, v in payload.human_mappings.items():
        if v:
            final_mappings[str(k)] = v 

    try:
        # 3. Download Flat File
        gcs.download_file(flat_path, local_flat)
        
        # 3. Re-process (Rename ONLY as rows are already cropped in flat)
        _process_excel_extract(
            local_flat, 
            local_validated, 
            header_row=1, # No more cropping needed
            mappings={int(k): v for k, v in final_mappings.items()}
        )
        
        # 4. Upload to GCS
        validated_blob = f"brief_media_files/{brief_id}/{plan.id}/validated-columns/plan_final.xlsx"
        gcs.upload_file(local_validated, validated_blob)
        
        # 5. Update DB
        plan.human_mappings = payload.human_mappings
        plan.validated_column_file_path = validated_blob
        plan.plan_file_url = validated_blob
        plan.plan_file_name = "plan_final.xlsx"
        plan.updated_at = models.get_utc_now()
        
        # 6. History
        history = models.HistoryTrail(
            agency_plan_id=plan.id,
            action="COLUMNS_UPDATED",
            user_id=current_user["id"],
            details="User confirmed/updated column mappings. Final file generated."
        )
        db.add(history)
        db.commit()
        
        # 7. Generate Signed URL
        final_url = gcs.get_signed_url(validated_blob, method="GET", expiration_minutes=8640)  # 6 days
        
        return {
            "validatedFileUrl": final_url,
            "status": "success"
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Update failed: {str(e)}")
    finally:
        if 'local_flat' in locals() and os.path.exists(local_flat): os.remove(local_flat)
        if os.path.exists(local_validated): os.remove(local_validated)

# Removed old endpoints: request_upload_url, confirm_upload, validate_rows

@app.get("/briefs/{brief_id}/plans/{plan_id}", response_model=plan_schema.AgencyPlanDetail)
def get_plan_detail(
    brief_id: int,
    plan_id: int,
    db: Session = Depends(session.get_db),
    current_user: dict = Depends(security.get_current_user)
):
    """
    Deep-dive into a specific plan (DS Group or Agency Owner).
    """
    plan = db.query(models.AgencyPlan).filter(
        models.AgencyPlan.id == plan_id,
        models.AgencyPlan.brief_id == brief_id
    ).first()

    print(plan)    
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")

    if current_user["role"] == "AGENCY" and plan.agency_id != current_user["agency_id"]:
        raise HTTPException(status_code=403, detail="Forbidden: You can only view your own plans.")

    # Generate Production VIEW URL if file exists
    view_url = None
    
    # if plan.validated_file_path:
    #     view_url = gcs.get_signed_url(plan.validated_file_path, method="GET")
    # elif plan.flat_file_path:
    #     view_url = gcs.get_signed_url(plan.flat_file_path, method="GET") # Fallback

    # User Request: Dynamic URL Fallback (Validated -> Flat -> Raw)
    latest_blob = plan.validated_column_file_path or plan.flat_file_path or plan.raw_file_path
    if latest_blob:
        final_plan_url = gcs.get_signed_url(latest_blob, method="GET", expiration_minutes=8640) # 6 days limit (GCS max is 7 days)
    else:
        final_plan_url = None

    final_plan_name = plan.plan_file_name or "Plan"

    return {
        "id": plan.id,
        "agencyId": plan.agency.id,
        "agencyName": plan.agency.name,
        "status": plan.status,
        "submittedAt": plan.submitted_at,
        "planFileName": final_plan_name,
        "planFileUrl": final_plan_url,
        "versionNumber": plan.version_number,
        "createdAt": plan.created_at,
        "updatedAt": plan.updated_at,
        "creator": plan.creator,
        "updater": plan.updater,
        "history": plan.history
    }

# --- VALIDATION STUBS ---

# --- VALIDATION & SUBMISSION ENDPOINTS ---

# validate columns and rows replaced by new flow

@app.post("/briefs/{brief_id}/submit")
def submit_plan(
    brief_id: int,
    payload: plan_schema.SubmitPlanRequest,
    db: Session = Depends(session.get_db),
    current_user: dict = Depends(security.verify_agency)
):
    """Officially submits the plan for Client Review."""
    plan = db.query(models.AgencyPlan).filter(
        models.AgencyPlan.brief_id == brief_id,
        models.AgencyPlan.agency_id == current_user["agency_id"]
    ).first()
    
    if not plan:
        raise HTTPException(status_code=404, detail="Plan slot not found")
        
    plan.status = "PENDING_REVIEW"
    plan.submitted_at = models.get_utc_now()
    plan.updated_at = models.get_utc_now()
    plan.updated_by = current_user["id"]
    
    # Audit History
    history = models.HistoryTrail(
        agency_plan_id=plan.id,
        action="PLAN_SUBMITTED",
        user_id=current_user["id"],
        details=payload.comment or "Plan submitted for review."
    )
    db.add(history)
    db.commit()
    
    return {"status": "success", "newStatus": plan.status}

@app.post("/briefs/{brief_id}/plans/{plan_id}/review", response_model=plan_schema.ReviewResponse)
def review_plan(
    brief_id: int,
    plan_id: int,
    review: plan_schema.ReviewSubmissionRequest,
    db: Session = Depends(session.get_db),
    current_user: dict = Depends(security.get_current_user)
):
    """
    Unified Review/Comment endpoint.
    - DS_GROUP: Can change status (APPROVE/REJECT) and add comments.
    - AGENCY: Can ONLY add comments (Status change forbidden), and only on THEIR OWN plan.
    """
    plan = db.query(models.AgencyPlan).filter(
        models.AgencyPlan.id == plan_id,
        models.AgencyPlan.brief_id == brief_id
    ).first()
    
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")

    # --- PERMISSION CHECK ---
    is_ds_group = current_user["role"] == "DS_GROUP"
    is_agency_owner = current_user["role"] == "AGENCY" and plan.agency_id == current_user["agency_id"]
    
    if not (is_ds_group or is_agency_owner):
        raise HTTPException(status_code=403, detail="Forbidden: You cannot review/comment on this plan.")

    # --- STATUS CHANGE CHECK ---
    has_status_change = review.status is not None
    
    # Agency cannot change status
    if not is_ds_group and has_status_change:
        raise HTTPException(status_code=403, detail="Forbidden: Agency users cannot change plan status.")
        
    old_status = plan.status
    
    if has_status_change:
        plan.status = review.status
        action_name = "STATUS_CHANGE"
        details_text = f"Status changed from {old_status} to {review.status}."
        history_comment = review.comment # Comment is optional with status change
    else:
        action_name = "COMMENT_ADDED"
        details_text = "A new comment was added to the plan history."
        history_comment = review.comment
    
    plan.updated_at = models.get_utc_now()
    plan.updated_by = current_user["id"]
    
    # Audit History
    history = models.HistoryTrail(
        agency_plan_id=plan.id,
        action=action_name,
        user_id=current_user["id"],
        details=details_text,
        comment=history_comment
    )
    db.add(history)
    
    # Update Brief Status ONLY if the plan status actually changed AND user is DS_GROUP (Redundant check but safe)
    brief = None
    if has_status_change and is_ds_group:
        brief = db.query(models.Brief).filter(models.Brief.id == brief_id).first()
        if brief:
            all_plans = db.query(models.AgencyPlan).filter(
                models.AgencyPlan.brief_id == brief_id
            ).all()
            
            approved_count = sum(1 for p in all_plans if p.status == "APPROVED")
            rejected_count = sum(1 for p in all_plans if p.status == "REJECTED")
            
            if approved_count == len(all_plans):
                brief.status = "APPROVED"
            elif rejected_count > 0:
                brief.status = "REJECTED"
            
            brief.updated_at = models.get_utc_now()
            brief.updated_by = current_user["id"]
    
    db.commit()
    db.refresh(history)
    
    return {
        "status": "success",
        "newPlanStatus": plan.status,
        "newBriefStatus": brief.status if brief else None,
        "history": history
    }
